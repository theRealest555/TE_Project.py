import pytest
from datetime import datetime
from fastapi.testclient import TestClient
import io
from app.security import create_access_token
from app.models import Submission, User, RoleType
from app.reports import ReportGenerator
from openpyxl import load_workbook


@pytest.fixture
def sample_submissions(db, regular_admin_user):
    """Create sample submissions for testing reports"""
    # Create a few test submissions
    submissions = [
        Submission(
            first_name="John",
            last_name="Doe",
            cin="AB123456",
            te_id="T12345",
            date_of_birth=datetime(1990, 1, 1),
            grey_card_number="12345-A-67890",
            plant="Plant1",
            cin_file_path="test/path/cin1.jpg",
            picture_file_path="test/path/pic1.jpg",
            grey_card_file_path="test/path/grey1.jpg",
            admin_id=regular_admin_user.id
        ),
        Submission(
            first_name="Jane",
            last_name="Smith",
            cin="CD789012",
            te_id="T67890",
            date_of_birth=datetime(1992, 5, 15),
            grey_card_number="67890-B-12345",
            plant="Plant1",
            cin_file_path="test/path/cin2.jpg",
            picture_file_path="test/path/pic2.jpg",
            grey_card_file_path="test/path/grey2.jpg",
            admin_id=regular_admin_user.id
        ),
        Submission(
            first_name="Michael",
            last_name="Johnson",
            cin="EF345678",
            te_id="T24680",
            date_of_birth=datetime(1985, 10, 20),
            grey_card_number="24680-C-13579",
            plant="Plant2",
            cin_file_path="test/path/cin3.jpg",
            picture_file_path="test/path/pic3.jpg",
            grey_card_file_path="test/path/grey3.jpg",
            admin_id=regular_admin_user.id
        )
    ]
    
    for submission in submissions:
        db.add(submission)
    
    db.commit()
    
    return submissions


def test_generate_format_1_report(db, sample_submissions):
    """Test generating Format 1 report"""
    # Generate report
    report_generator = ReportGenerator()
    output, filename = report_generator.generate_format_1(db)
    
    # Load the Excel file from the BytesIO object
    wb = load_workbook(output)
    ws = wb.active
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Last Name"
    assert ws.cell(row=1, column=2).value == "First Name"
    assert ws.cell(row=1, column=3).value == "CIN"
    assert ws.cell(row=1, column=4).value == "TE ID"
    assert ws.cell(row=1, column=5).value == "Date of Birth"
    
    # Check data (all submissions should be included)
    assert ws.max_row == 4  # Header + 3 submissions
    
    # Check first submission
    assert ws.cell(row=2, column=1).value == "Doe"
    assert ws.cell(row=2, column=2).value == "John"
    assert ws.cell(row=2, column=3).value == "AB123456"
    assert ws.cell(row=2, column=4).value == "T12345"
    assert ws.cell(row=2, column=5).value == "1990-01-01"


def test_generate_format_2_report(db, sample_submissions):
    """Test generating Format 2 report"""
    # Generate report
    report_generator = ReportGenerator()
    output, filename = report_generator.generate_format_2(db)
    
    # Load the Excel file from the BytesIO object
    wb = load_workbook(output)
    ws = wb.active
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Last Name"
    assert ws.cell(row=1, column=2).value == "First Name"
    assert ws.cell(row=1, column=3).value == "Grey Card Number"
    assert ws.cell(row=1, column=4).value == "TE ID"
    
    # Check data (all submissions should be included)
    assert ws.max_row == 4  # Header + 3 submissions
    
    # Check first submission
    assert ws.cell(row=2, column=1).value == "Doe"
    assert ws.cell(row=2, column=2).value == "John"
    assert ws.cell(row=2, column=3).value == "12345-A-67890"
    assert ws.cell(row=2, column=4).value == "T12345"


def test_generate_report_with_plant_filter(db, sample_submissions):
    """Test generating report with plant filter"""
    # Generate report for Plant1 only
    report_generator = ReportGenerator()
    output, filename = report_generator.generate_format_1(db, plant="Plant1")
    
    # Load the Excel file from the BytesIO object
    wb = load_workbook(output)
    ws = wb.active
    
    # Check data (only Plant1 submissions should be included)
    assert ws.max_row == 3  # Header + 2 submissions
    
    # Check that all rows are from Plant1
    plant1_names = ["Doe", "Smith"]
    for row in range(2, ws.max_row + 1):
        assert ws.cell(row=row, column=1).value in plant1_names


def test_reports_api_endpoint_format_1(client, super_admin_user, sample_submissions):
    """Test the report API endpoint for format 1"""
    # Create access token for super admin
    access_token = create_access_token(data={"sub": super_admin_user.username})
    
    # Get format 1 report
    response = client.get(
        "/admin/reports?report_format=1",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type".lower()] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment; filename=" in response.headers["content-disposition".lower()]
    assert ".xlsx" in response.headers["content-disposition".lower()]
    
    # Read the Excel file from the response content
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Last Name"
    assert ws.cell(row=1, column=5).value == "Date of Birth"


def test_reports_api_endpoint_format_2(client, super_admin_user, sample_submissions):
    """Test the report API endpoint for format 2"""
    # Create access token for super admin
    access_token = create_access_token(data={"sub": super_admin_user.username})
    
    # Get format 2 report
    response = client.get(
        "/admin/reports?report_format=2",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type".lower()] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Read the Excel file from the response content
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check headers
    assert ws.cell(row=1, column=1).value == "Last Name"
    assert ws.cell(row=1, column=3).value == "Grey Card Number"


def test_regular_admin_can_only_see_own_plant_data(client, regular_admin_user, sample_submissions):
    """Test that regular admin can only generate reports for their own plant"""
    # Create access token for regular admin
    access_token = create_access_token(data={"sub": regular_admin_user.username})
    
    # Get format 1 report
    response = client.get(
        "/admin/reports?report_format=1",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 200
    
    # Read the Excel file from the response content
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check data (only Plant1 submissions should be included since regular_admin is from Plant1)
    assert ws.max_row == 3  # Header + 2 submissions
    
    # Check that all rows are from Plant1
    plant1_names = ["Doe", "Smith"]
    for row in range(2, ws.max_row + 1):
        assert ws.cell(row=row, column=1).value in plant1_names


def test_invalid_report_format(client, super_admin_user):
    """Test that invalid report format returns validation error"""
    # Create access token for super admin
    access_token = create_access_token(data={"sub": super_admin_user.username})
    
    # Try to generate report with invalid format
    response = client.get(
        "/admin/reports?report_format=3",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 422  # Validation error


def test_unauthorized_report_access(client):
    """Test that unauthenticated access is denied"""
    # Try to generate report without authentication
    response = client.get("/admin/reports?report_format=1")
    
    assert response.status_code == 401
    assert response.json()["detail"] == "Not authenticated"


def test_report_with_no_data(client, super_admin_user, db):
    """Test generating report when no data exists"""
    # Clear any existing submissions
    db.query(Submission).delete()
    db.commit()
    
    # Create access token for super admin
    access_token = create_access_token(data={"sub": super_admin_user.username})
    
    # Generate report format 1 with no data in database
    response = client.get(
        "/admin/reports?report_format=1",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type".lower()] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "content-disposition" in [k.lower() for k in response.headers.keys()]
    assert "attachment; filename=employee_data_format1_" in response.headers["content-disposition".lower()]


def test_report_with_plant_filter_api(client, super_admin_user, sample_submissions):
    """Test the API endpoint with plant filter"""
    # Create access token for super admin
    access_token = create_access_token(data={"sub": super_admin_user.username})
    
    # Generate report format 1 with plant filter
    response = client.get(
        "/admin/reports?report_format=1&plant=Plant1",
        headers={"Authorization": f"Bearer {access_token}"}
    )
    
    assert response.status_code == 200
    assert response.headers["content-type".lower()] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "content-disposition" in [k.lower() for k in response.headers.keys()]
    
    # Read the Excel file from the response content
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    
    # Check data (only Plant1 submissions should be included)
    assert ws.max_row == 3  # Header + 2 submissions
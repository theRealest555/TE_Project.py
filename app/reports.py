from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from .models import Submission
from io import BytesIO
from datetime import datetime


class ReportGenerator:
    @staticmethod
    def generate_format_1(db: Session, plant: str = None):
        """
        Generate Format 1 report: Last Name, First Name, CIN, TE ID, Date of Birth
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Data"
        
        # Set up headers
        headers = ["Last Name", "First Name", "CIN", "TE ID", "Date of Birth"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Query data
        query = db.query(Submission)
        if plant:
            query = query.filter(Submission.plant == plant)
        submissions = query.all()
        
        # Add data rows
        for row_num, submission in enumerate(submissions, 2):
            ws.cell(row=row_num, column=1).value = submission.last_name
            ws.cell(row=row_num, column=2).value = submission.first_name
            ws.cell(row=row_num, column=3).value = submission.cin
            ws.cell(row=row_num, column=4).value = submission.te_id
            ws.cell(row=row_num, column=5).value = submission.date_of_birth.strftime("%Y-%m-%d")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, f"employee_data_format1_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    @staticmethod
    def generate_format_2(db: Session, plant: str = None):
        """
        Generate Format 2 report: Last Name, First Name, Grey Card Number, TE ID
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Grey Cards"
        
        # Set up headers
        headers = ["Last Name", "First Name", "Grey Card Number", "TE ID"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Query data
        query = db.query(Submission)
        if plant:
            query = query.filter(Submission.plant == plant)
        submissions = query.all()
        
        # Add data rows
        for row_num, submission in enumerate(submissions, 2):
            ws.cell(row=row_num, column=1).value = submission.last_name
            ws.cell(row=row_num, column=2).value = submission.first_name
            ws.cell(row=row_num, column=3).value = submission.grey_card_number
            ws.cell(row=row_num, column=4).value = submission.te_id
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, f"employee_grey_cards_format2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


report_generator = ReportGenerator()